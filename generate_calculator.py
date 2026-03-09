from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "kredyt_vs_wynajem_mvp.xlsx"
MAX_MONTHS = 360
SALES_POINTS_ROWS = 12


def style_headers(ws, header_row: int = 1, start_col: int = 1, end_col: int | None = None) -> None:
    if end_col is None:
        end_col = ws.max_column
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = fill
        cell.font = font


def build_assumptions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Założenia"

    ws["A1"] = "Kalkulator: Kredyt vs Wynajem (MVP)"
    ws["A1"].font = Font(size=14, bold=True)

    labels = [
        ("A2", "Wartość nieruchomości"),
        ("A3", "Wkład własny"),
        ("A4", "Kwota kredytu"),
        ("A5", "LTV (%) - steruje kwotą kredytu"),
        ("A6", "Okres spłaty (miesiące)"),
        ("A7", "Informacja o latach"),
        ("A8", "Marża"),
        ("A9", "WIBOR niski"),
        ("A10", "WIBOR bazowy"),
        ("A11", "WIBOR wysoki"),
        ("A12", "Scenariusz aktywny (podgląd)"),
        ("A13", "Tryb nadpłaty"),
        ("A14", "Stała nadpłata miesięczna"),
        ("A15", "Roczna zmiana wartości mieszkania"),
        ("A16", "Koszt wejścia w zakup (%)"),
        ("A17", "Koszt sprzedaży (%)"),
        ("A18", "Najem miesięczny startowy"),
        ("A19", "Model zmiany najmu"),
        ("A20", "Skok najmu co 24 miesiące"),
        ("A21", "Inwestowanie różnicy"),
        ("A22", "Stopa inwestycji różnicy"),
        ("A23", "Kapitał startowy najemcy"),
    ]

    for addr, label in labels:
        ws[addr] = label

    ws["B2"] = 700000
    ws["B5"] = 0.80
    ws["B4"] = "=B2*B5"
    ws["B3"] = "=B2-B4"
    ws["B6"] = 360
    ws["B7"] = '=IF(MOD(B6,12)=0,B6/12&" lat","")'
    ws["B8"] = 0.018
    ws["B9"] = 0.025
    ws["B10"] = 0.04
    ws["B11"] = 0.06
    ws["B12"] = "Bazowy"
    ws["B13"] = "Stała"
    ws["B14"] = 0
    ws["B15"] = 0.03
    ws["B16"] = 0.03
    ws["B17"] = 0.02
    ws["B18"] = 3000
    ws["B19"] = "Skokowy"
    ws["B20"] = 500
    ws["B21"] = "ON"
    ws["B22"] = "=B10"
    ws["B23"] = "=B3"

    ws["D1"] = "Miesiące sprzedaży"
    ws["D2"] = 24
    ws["D3"] = 36
    ws["D4"] = 48

    style_headers(ws, 1, 1, 4)

    percent_cells = ["B5", "B8", "B9", "B10", "B11", "B15", "B16", "B17", "B22"]
    for cell in percent_cells:
        ws[cell].number_format = "0.00%"

    currency_cells = ["B2", "B3", "B4", "B14", "B18", "B20", "B23"]
    for cell in currency_cells:
        ws[cell].number_format = "#,##0.00"

    dv_scenario = DataValidation(type="list", formula1='"Niski,Bazowy,Wysoki"', allow_blank=False)
    ws.add_data_validation(dv_scenario)
    dv_scenario.add("B12")

    dv_overpay = DataValidation(type="list", formula1='"Stała,Harmonogram"', allow_blank=False)
    ws.add_data_validation(dv_overpay)
    dv_overpay.add("B13")

    dv_rent_model = DataValidation(type="list", formula1='"Stały,Skokowy"', allow_blank=False)
    ws.add_data_validation(dv_rent_model)
    dv_rent_model.add("B19")

    dv_invest = DataValidation(type="list", formula1='"ON,OFF"', allow_blank=False)
    ws.add_data_validation(dv_invest)
    dv_invest.add("B21")

    ws["E5"] = "Przykład: 80% oznacza kredyt na 80% wartości mieszkania"
    ws["E20"] = "Skok o tyle PLN co każde 24 miesiące, gdy model = Skokowy"
    ws["E22"] = "Domyślnie = WIBOR bazowy, możesz nadpisać ręcznie"

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 62


def build_amortization_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Amortyzacja")

    headers = [
        "Miesiąc",
        "Saldo start (Niski)",
        "Rata (Niski)",
        "Odsetki (Niski)",
        "Kapitał (Niski)",
        "Nadpłata (Niski)",
        "Saldo koniec (Niski)",
        "Saldo start (Bazowy)",
        "Rata (Bazowy)",
        "Odsetki (Bazowy)",
        "Kapitał (Bazowy)",
        "Nadpłata (Bazowy)",
        "Saldo koniec (Bazowy)",
        "Saldo start (Wysoki)",
        "Rata (Wysoki)",
        "Odsetki (Wysoki)",
        "Kapitał (Wysoki)",
        "Nadpłata (Wysoki)",
        "Saldo koniec (Wysoki)",
        "Nadpłata harmonogram (input)",
        "Nadpłata efektywna",
    ]

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    style_headers(ws, 1, 1, len(headers))

    for row in range(2, MAX_MONTHS + 2):
        ws[f"A{row}"] = row - 1

        ws[f"U{row}"] = f'=IF(Założenia!$B$13="Harmonogram",MAX($T{row},0),MAX(Założenia!$B$14,0))'

        if row == 2:
            ws[f"B{row}"] = "=Założenia!$B$4"
        else:
            ws[f"B{row}"] = f"=G{row - 1}"
        ws[f"C{row}"] = f"=IF($A{row}<=Założenia!$B$6,PMT((Założenia!$B$8+Założenia!$B$9)/12,Założenia!$B$6,-Założenia!$B$4),0)"
        ws[f"D{row}"] = f"=IF(B{row}<=0,0,B{row}*(Założenia!$B$8+Założenia!$B$9)/12)"
        ws[f"E{row}"] = f"=IF(B{row}<=0,0,MIN(C{row}-D{row},B{row}))"
        ws[f"F{row}"] = f"=IF(B{row}<=0,0,MIN($U{row},B{row}-E{row}))"
        ws[f"G{row}"] = f"=MAX(B{row}-E{row}-F{row},0)"

        if row == 2:
            ws[f"H{row}"] = "=Założenia!$B$4"
        else:
            ws[f"H{row}"] = f"=M{row - 1}"
        ws[f"I{row}"] = f"=IF($A{row}<=Założenia!$B$6,PMT((Założenia!$B$8+Założenia!$B$10)/12,Założenia!$B$6,-Założenia!$B$4),0)"
        ws[f"J{row}"] = f"=IF(H{row}<=0,0,H{row}*(Założenia!$B$8+Założenia!$B$10)/12)"
        ws[f"K{row}"] = f"=IF(H{row}<=0,0,MIN(I{row}-J{row},H{row}))"
        ws[f"L{row}"] = f"=IF(H{row}<=0,0,MIN($U{row},H{row}-K{row}))"
        ws[f"M{row}"] = f"=MAX(H{row}-K{row}-L{row},0)"

        if row == 2:
            ws[f"N{row}"] = "=Założenia!$B$4"
        else:
            ws[f"N{row}"] = f"=S{row - 1}"
        ws[f"O{row}"] = f"=IF($A{row}<=Założenia!$B$6,PMT((Założenia!$B$8+Założenia!$B$11)/12,Założenia!$B$6,-Założenia!$B$4),0)"
        ws[f"P{row}"] = f"=IF(N{row}<=0,0,N{row}*(Założenia!$B$8+Założenia!$B$11)/12)"
        ws[f"Q{row}"] = f"=IF(N{row}<=0,0,MIN(O{row}-P{row},N{row}))"
        ws[f"R{row}"] = f"=IF(N{row}<=0,0,MIN($U{row},N{row}-Q{row}))"
        ws[f"S{row}"] = f"=MAX(N{row}-Q{row}-R{row},0)"

    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U"]:
        for row in range(2, MAX_MONTHS + 2):
            ws[f"{col}{row}"].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 10
    for col in "BCDEFGHIJKLMNOPQRSTU":
        ws.column_dimensions[col].width = 18


def build_rent_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Wynajem")

    headers = [
        "Miesiąc",
        "Najem miesięczny",
        "Skumulowany koszt najmu",
        "Koszt kredytu (scenariusz aktywny)",
        "Różnica do inwestycji",
        "Inwestycja różnic",
        "Majątek najemcy netto",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    style_headers(ws, 1, 1, len(headers))

    for row in range(2, MAX_MONTHS + 2):
        ws[f"A{row}"] = row - 1
        ws[f"B{row}"] = f"=IF(Założenia!$B$19=\"Stały\",Założenia!$B$18,Założenia!$B$18+INT((A{row}-1)/24)*Założenia!$B$20)"
        if row == 2:
            ws[f"C{row}"] = f"=B{row}"
        else:
            ws[f"C{row}"] = f"=C{row - 1}+B{row}"

        ws[f"D{row}"] = (
            f"=IF(Założenia!$B$12=\"Niski\",Amortyzacja!C{row}+Amortyzacja!F{row},"
            f"IF(Założenia!$B$12=\"Bazowy\",Amortyzacja!I{row}+Amortyzacja!L{row},"
            f"Amortyzacja!O{row}+Amortyzacja!R{row}))"
        )
        ws[f"E{row}"] = f"=MAX(D{row}-B{row},0)"

        if row == 2:
            ws[f"F{row}"] = f"=IF(Założenia!$B$21=\"ON\",E{row}*(1+Założenia!$B$22/12),E{row})"
        else:
            ws[f"F{row}"] = f"=IF(Założenia!$B$21=\"ON\",F{row-1}*(1+Założenia!$B$22/12)+E{row},F{row-1}+E{row})"

        ws[f"G{row}"] = f"=Założenia!$B$23+F{row}-C{row}"

    for col in "BCDEFG":
        for row in range(2, MAX_MONTHS + 2):
            ws[f"{col}{row}"].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 10
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 30


def build_comparison_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Porównanie")

    headers = ["miesiąc_sprzedaży", "scenariusz_wibor", "majątek_kupno", "majątek_wynajem", "różnica"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    style_headers(ws, 1, 1, 5)

    scenario_names = ["Niski", "Bazowy", "Wysoki"]

    row = 2
    for idx in range(SALES_POINTS_ROWS):
        src_row = idx + 2
        for scenario in scenario_names:
            ws[f"A{row}"] = f"=IF(OR(Założenia!$D${src_row}=\"\",Założenia!$D${src_row}>Założenia!$B$6),\"\",Założenia!$D${src_row})"
            ws[f"B{row}"] = scenario

            if scenario == "Niski":
                saldo_range = "Amortyzacja!$G$2:$G$361"
            elif scenario == "Bazowy":
                saldo_range = "Amortyzacja!$M$2:$M$361"
            else:
                saldo_range = "Amortyzacja!$S$2:$S$361"

            ws[f"C{row}"] = (
                f"=IF($A{row}=\"\",\"\","
                f"Założenia!$B$2*(1+Założenia!$B$15)^($A{row}/12)*(1-Założenia!$B$17)"
                f"-INDEX({saldo_range},$A{row})"
                f"-Założenia!$B$2*Założenia!$B$16)"
            )
            ws[f"D{row}"] = f"=IF($A{row}=\"\",\"\",INDEX(Wynajem!$G$2:$G$361,$A{row}))"
            ws[f"E{row}"] = f"=IF($A{row}=\"\",\"\",C{row}-D{row})"
            row += 1

    for col in ["C", "D", "E"]:
        for row in range(2, 2 + SALES_POINTS_ROWS * 3):
            ws[f"{col}{row}"].number_format = "#,##0.00"

    ws["H1"] = "miesiąc"
    ws["I1"] = "Kupno Niski"
    ws["J1"] = "Kupno Bazowy"
    ws["K1"] = "Kupno Wysoki"
    ws["L1"] = "Wynajem"
    style_headers(ws, 1, 8, 12)

    for i in range(SALES_POINTS_ROWS):
        r = i + 2
        ws[f"H{r}"] = f"=Założenia!$D${r}"

        ws[f"I{r}"] = (
            f"=IF(OR($H{r}=\"\",$H{r}>Założenia!$B$6),NA(),"
            f"Założenia!$B$2*(1+Założenia!$B$15)^($H{r}/12)*(1-Założenia!$B$17)"
            f"-INDEX(Amortyzacja!$G$2:$G$361,$H{r})-Założenia!$B$2*Założenia!$B$16)"
        )
        ws[f"J{r}"] = (
            f"=IF(OR($H{r}=\"\",$H{r}>Założenia!$B$6),NA(),"
            f"Założenia!$B$2*(1+Założenia!$B$15)^($H{r}/12)*(1-Założenia!$B$17)"
            f"-INDEX(Amortyzacja!$M$2:$M$361,$H{r})-Założenia!$B$2*Założenia!$B$16)"
        )
        ws[f"K{r}"] = (
            f"=IF(OR($H{r}=\"\",$H{r}>Założenia!$B$6),NA(),"
            f"Założenia!$B$2*(1+Założenia!$B$15)^($H{r}/12)*(1-Założenia!$B$17)"
            f"-INDEX(Amortyzacja!$S$2:$S$361,$H{r})-Założenia!$B$2*Założenia!$B$16)"
        )
        ws[f"L{r}"] = f"=IF(OR($H{r}=\"\",$H{r}>Założenia!$B$6),NA(),INDEX(Wynajem!$G$2:$G$361,$H{r}))"

    for col in "IJKL":
        for row in range(2, 2 + SALES_POINTS_ROWS):
            ws[f"{col}{row}"].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["H"].width = 12
    for col in "IJKL":
        ws.column_dimensions[col].width = 18


def add_scenario_chart(target_ws, src_ws, title: str, buy_col: int, anchor: str) -> None:
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Miesiąc sprzedaży"
    chart.y_axis.title = "Majątek netto [PLN]"
    chart.height = 8
    chart.width = 19
    chart.legend.position = "b"

    xvalues = Reference(src_ws, min_col=8, min_row=2, max_row=2 + SALES_POINTS_ROWS - 1)
    y_buy = Reference(src_ws, min_col=buy_col, min_row=1, max_row=2 + SALES_POINTS_ROWS - 1)
    y_rent = Reference(src_ws, min_col=12, min_row=1, max_row=2 + SALES_POINTS_ROWS - 1)

    s1 = Series(y_buy, xvalues, title_from_data=True)
    s1.graphicalProperties.line.width = 22000
    s1.marker.symbol = "circle"
    s1.marker.size = 6

    s2 = Series(y_rent, xvalues, title_from_data=True)
    s2.graphicalProperties.line.width = 22000
    s2.marker.symbol = "triangle"
    s2.marker.size = 6

    chart.series.append(s1)
    chart.series.append(s2)

    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"

    target_ws.add_chart(chart, anchor)


def build_chart_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Wykres")
    ws["A1"] = "Porównanie majątku netto po sprzedaży"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "3 wykresy: każdy scenariusz WIBOR osobno, wspólna linia Wynajmu"

    src_ws = wb["Porównanie"]
    add_scenario_chart(ws, src_ws, "Scenariusz WIBOR Niski", 9, "A4")
    add_scenario_chart(ws, src_ws, "Scenariusz WIBOR Bazowy", 10, "A23")
    add_scenario_chart(ws, src_ws, "Scenariusz WIBOR Wysoki", 11, "A42")


def main() -> None:
    wb = Workbook()
    build_assumptions_sheet(wb)
    build_amortization_sheet(wb)
    build_rent_sheet(wb)
    build_comparison_sheet(wb)
    build_chart_sheet(wb)

    out_path = Path(OUTPUT_FILE)
    wb.save(out_path)
    print(f"Workbook generated: {out_path.resolve()}")


if __name__ == "__main__":
    main()
